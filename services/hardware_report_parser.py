import os
import openpyxl


def get_path_to_excel_file(directory):
    """Получаем список путей до файлов"""
    path_list = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            path_list.append(root + '\\' + filename)
    # print(path_list)
    return path_list


def get_list_from_hardware_report(path, sheet_name):
    """Получаем список с контентом из файла"""
    # print(path)
    path_list = path.split('\\')
    # print(len(path_list), path_list)
    # Получаем имя проекта из имени файла или из имени каталога
    if len(path_list) == 5:
        project = path_list[3]
    else:
        project = path_list[3][:-5]
    cell_content_list = []
    work_book_of_excel_file = openpyxl.load_workbook(path)
    sheet_of_work_book = work_book_of_excel_file[sheet_name]
    for cell in sheet_of_work_book['A']:
        if cell.value is not None:
            cell_content_str = str(cell.value)
            cell_content_list.append(cell_content_str.split())
    cell_content_list.append(['root_project', project])
    return cell_content_list


def get_list_of_lists_from_cell_content_list(cell_content_list):
    """Формируем список списков из списка контента"""
    rack_index = 0
    rack_list = []
    # Находим Rack, запоминаем индекс, находим следующее упоминание Rack, формируем список, добавляем в список.
    for index in range(len(cell_content_list)):
        rack_index_start = rack_index
        # print(cell_content_list[index])
        if ('Rack' in cell_content_list[index] or 'Rack:' in cell_content_list[index]) \
                and 'Loss' not in cell_content_list[index]:
            rack_index = index
            # print(rack_index_start, rack_index)
            rack_list.append(cell_content_list[rack_index_start:rack_index])
    rack_list.append(cell_content_list[rack_index:])
    return rack_list


def get_hardware_report_dict(rack_list):
    """Формируем словарь с хардварным репортом"""
    hardware_report_dict = {
        'name': '',
        'ip': [],
        'racks': []
    }
    for item in rack_list:
        # print(item)
        for element in item:
            # print(element)
            if 'root_project' in element:
                hardware_report_dict['project'] = element[1]
            if 'Name:' in element and 'Target' in element:
                plc_name = element[2]
                hardware_report_dict['name'] = plc_name
            if 'IP' in element and 'Address:' in element and len(element) == 3:
                ip = element[2]
                hardware_report_dict['ip'].append(element[2])
            if 'Rack' in element or 'Rack:' in element:
                # print(element)
                if 'Loss' not in element:
                    # print(element)
                    if element[2] != 'Slot':
                        # print(element)
                        rack_model = element[2][:-1]
                        rack_type = ' '.join(element[3:])
                        rack_number = element[1][:-1]
                        if rack_number == 'Rack':
                            rack_number = '0'
                        hardware_report_dict['racks'].append(
                            {'number': rack_number, 'type': rack_type, 'model': rack_model, 'slots': []})
            if ('Rack' in element or 'Rack:' in element) and 'Slot' in element:
                # print(element)
                rack_number = element[1][:-1]
                if rack_number == 'Rack':
                    rack_number = '0'
                if rack_number == '':
                    rack_number = element[1]
                slot_number = element[3]
                if 'Used' not in item[1] and '3RD' not in item[1]:
                    # print(item[1])
                    slot_model = item[1][0][:-1]
                    # print(slot_model[0:2])
                    if slot_model[0:2] != 'IC':
                        slot_model = ' '.join(item[1][0:3])
                        slot_model = slot_model[:-1]
                        slot_type = ' '.join(item[1][3:])
                        if slot_model[0:2] == 'HE':
                            slot_model = slot_model[:-1]
                            slot_type = ' '.join(item[1][1:])
                        #     print(slot_model)
                        #     print(slot_type)
                        # print(slot_model, slot_type)
                    slot_type = ' '.join(item[1][2:])
                    # print(slot_type)
                    if item[1][0][-1] != ',':
                        # print(item)
                        slot_model = item[1][0]
                    # print(slot_model, slot_type)
                    hardware_report_dict['racks'][int(rack_number)]['slots'].append(
                        {'number': slot_number, 'type': slot_type, 'model': slot_model, 'reference_address': []})
                    index_slots = len(hardware_report_dict['racks'][int(rack_number)]['slots']) - 1
                    for el in item:
                        reference_address = []
                        # print(len(hardware_report_dict['racks'][int(rack_number)]['slots'][index_slots][
                        #                 'reference_address']))
                        index_reference_address = len(
                            hardware_report_dict['racks'][int(rack_number)]['slots'][index_slots][
                                'reference_address'])
                        if 'Reference' in el and 'Address:' in el:
                            index_rack_list = item.index(el)
                            if 'Value' in el or 'Status' in el or 'Input' in el or 'Output' in el or 'Diagnostic' \
                                    in el or 'Outputs' in el:
                                address = item[index_rack_list][-1]
                            else:
                                address = item[index_rack_list][2]
                            # print(el)
                            # print(item.index(el))

                            length = item[index_rack_list + 1][1]
                            if length == 'Value' or length == 'Status' or length == 'Diagnostic' or \
                                    length == 'Reference' or length == 'Channel':
                                length = item[index_rack_list + 1][-1]
                            # print(item[index_rack_list], item[index_rack_list + 1], length, address)
                            hardware_report_dict['racks'][int(rack_number)]['slots'][index_slots][
                                'reference_address'].append({'address': address, 'length': length})
    return hardware_report_dict


def get_address_of_proscon_point(path):
    """Получаем из файла c Proscon переменными наименование точки, её адресс и описание"""
    point_dict = {'point': []}
    work_book_of_excel_file = openpyxl.load_workbook(path)
    sheet_of_work_book = work_book_of_excel_file.active
    for row in sheet_of_work_book.iter_rows():
        for index in range(len(row)):
            # print(row[index].value)
            if row[index].value is not None:
                cell_content = row[index].value
                cell_content_str = str(cell_content)
                # print(cell_content_str[0][0])
                if cell_content_str[0] == '%' and len(cell_content_str) > 1:
                    # print(cell_content_str, row[0].value)
                    point_dict['point'].append(
                        {'point_id': row[0].value, 'address': cell_content_str, 'description': row[1].value})
    return point_dict


def get_point_address(path):
    point_dict = {'point': []}
    work_book_of_excel_file = openpyxl.load_workbook(path)
    sheet_of_work_book = work_book_of_excel_file.active
    for row in sheet_of_work_book.iter_rows():
        if row[15].value is not None and row[15].value[0] == '%':
            description = row[2].value
            point_id = row[0].value
            address = row[15].value
            # print(path.split('\\'))
            plc_name = None
            project = path.split('\\')[-1][:path.split('\\')[-1].index('.')]
            if len(path.split('\\')) == 5:
                project = path.split('\\')[3]
                plc_name = path.split('\\')[-1][:path.split('\\')[-1].index('.')]
            # print(project)
            # print(point_id, description, address, project, plc_name)
            point_dict['point'].append(
                {'point_id': point_id, 'address': address, 'description': description, 'project': project, 'plc_name': plc_name})
    return point_dict
# for path in get_path_to_excel_file('D:\\config plc'):
#     cell_content_list = get_list_from_hardware_report(path)
#     rack_list = get_list_of_lists_from_cell_content_list(cell_content_list)
#     hardware_report_dict = get_hardware_report_dict(rack_list)
#     print(hardware_report_dict)
